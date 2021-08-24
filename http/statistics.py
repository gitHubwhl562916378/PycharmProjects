import re
from openpyxl import  Workbook 
from datetime import datetime

def split_camera_ability(file_name):
    out_data = {}
    with open(file_name) as f:
        iter_f = iter(f)
        for line in iter_f:
            patten = re.compile(r'\[(.+)\] \[multi_sink\].+ObjectDetectCommand\]: (.+) (.+) predict use (.+)$')
            match = patten.match(line)
            if not match:
                continue
            
            time = datetime.strptime(match.group(1),'%Y-%m-%d %H:%M:%S.%f')
            camera_id = match.group(2)
            ability_name = match.group(3)
            detect_use = float(match.group(4))
            if camera_id in out_data:
                camera_map = out_data[camera_id]
                if ability_name in camera_map:
                    camera_map[ability_name].append({"time": time, "detect_use": detect_use})
                else:
                    camera_map[ability_name] = [{"time": time, "detect_use": detect_use}]
                out_data[camera_id] = camera_map
            else:
                out_data[camera_id] = {ability_name: [{"time": time, "detect_use": detect_use}]}
    return out_data

def split_alarm_count(file_name):
    out_data = {}
    with open(file_name) as f:
        iter_f = iter(f)
        for line in iter_f:
            patten = re.compile(r'\[(.+)\] \[multi_sink\].+MagicMessageNotifier\]: (.+) HandleMessage (.+) post to \S+$')
            match = patten.match(line)
            if not match:
                continue
            
            time = datetime.strptime(match.group(1),'%Y-%m-%d %H:%M:%S.%f')
            camera_id = match.group(2)
            ability_name = match.group(3)
            if camera_id in out_data:
                camera_map = out_data[camera_id]
                if ability_name in camera_map:
                    camera_map[ability_name].append(time)
                else:
                    camera_map[ability_name] = [time]
                out_data[camera_id] = camera_map
            else:
                out_data[camera_id] = {ability_name: [time]}
    
    return out_data
if __name__ == '__main__':
    ab_name_map = {
    "RoadDamaged":"道路破损",
    "HumanBreakIn":"人员闯入",
    "OffStoreOperation":"店外经营",
    "GarbageExposed":"暴露垃圾",
    "NoneMotorVehicleParking":"非机动车乱停放",
    "MotorVehicleParking":"机动车乱停放",
    "BannerOrSlogansHungging":"悬挂标语横幅",
    "WithoutHelmetOnSite":"工地不戴安全帽",
    "VehicleLicenseRecon":"车牌识别",
    "FireOrSmokeDetected":"烟火智能感知",
    "FollowIntoHouseMotorbike":"电动车检测",
    "FollowIntoHousePerson":"人员尾随入户", 
    "RoadOccupatedInConstruction":"施工占道",
    "ResidueAccumulation":"积存垃圾渣土",  
    "VehicleOfDangerous":"危险车辆",
    "SocialVehicleDetect":"社会车辆检测",
    "IllegalAnUmbrella":"乱搭棚伞",
    "HangDownTheStreet":"沿街晾挂",
    "UnexpectedEvents":"人群聚集",
    "wellCoverAbnormal":"井盖异常",
    "UnlicensedBusinessVendor":"无照经营游商",
    "RoadPonding":"道路积水",
    "GarbageOverflow": "垃圾满溢",
    "RoadOccupatedInOperation": "占道经营"
    }
    wb = Workbook()
    sheet = wb.active
    split_data = split_camera_ability("logs/sdk.log")
    split_alarm = split_alarm_count("logs/sdk.log")
    
    cell_row = 1
    label_list = ["摄像头id", "AI能力", "报警数量","总共检测次数", "检测平均耗时", "平均检测间隔", "最高检测间隔", "大于3秒的检测间隔次数"]
    for col, label in enumerate(label_list, 1):
        sheet.cell(row=cell_row, column=col,value=label)
    cell_row+=1
    for camera_id in split_data:
        camera_data = split_data[camera_id]
        for ability in camera_data:
            ability_data_array = camera_data[ability]
            sum_use = 0.0
            sum_time_interval = 0.0
            highest_value = 0.0
            greter_count = 0
            last_index_data = ability_data_array[0]["time"]
            for index,ele_data in enumerate(ability_data_array):
                sum_use += ele_data["detect_use"]
                if ele_data["time"] == last_index_data:
                    continue
                else:
                    time_interval = ele_data["time"].timestamp() * 1000 - last_index_data.timestamp() * 1000
                    if time_interval > highest_value:
                        highest_value = time_interval
                    if time_interval > 1504 and time_interval < 1506:
                        print(camera_id, ability, ability_data_array[index-1]["time"], ele_data["time"])
                    if time_interval > 3000:
                        greter_count += 1
                    sum_time_interval += time_interval
                    last_index_data = ele_data["time"]
            sheet.cell(row=cell_row, column=1, value=camera_id)
            sheet.cell(row=cell_row, column=2, value=ab_name_map[ability])
            sheet.cell(row=cell_row, column=3, value=len(split_alarm[camera_id][ability]))
            sheet.cell(row=cell_row, column=4, value=len(ability_data_array))
            sheet.cell(row=cell_row, column=5, value=sum_use/len(ability_data_array))
            sheet.cell(row=cell_row, column=6, value=sum_time_interval/(len(ability_data_array)-1))
            sheet.cell(row=cell_row, column=7, value=highest_value)
            sheet.cell(row=cell_row, column=8, value=greter_count)
            cell_row += 1
    
    wb.save("results.xlsx")
