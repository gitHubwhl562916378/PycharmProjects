#!/usr/bin/python3.5
# -*- coding: UTF-8 -*-

# from time import *
# from datetime import datetime
#
# start_time = datetime.strptime('2050-04-02 15:56:00', '%Y-%m-%d %H:%M:%S')
# end_time = datetime.strptime('2050-04-02 15:57:10', '%Y-%m-%d %H:%M:%S')
# print((end_time - start_time).seconds)
#
# s = {1,2}
# s.add(8)
# print(s)
#'2020-04-03 13:02:02.812 [Debug]: HandAlgorithmTask::execute statisc camera[%s] face_body_count[%d] process face[%d]   recv_face_feature[%d]; process body[%d] recv_body_feature[%d]'
import subprocess
import re
from openpyxl import Workbook
import json

# with open('cameara', 'r') as f:
#     load_dict = json.load(f)
#     cameras = load_dict['cameras']
#     print('camera size: %d'%(len(cameras)))
#     for obj in cameras:
#         print(obj['id'], obj['rtsp'])


fh = subprocess.Popen("tail -n 100000 /home/whl/seye-logs/KLMediaServer.log", stdout=subprocess.PIPE, shell=True)
arr = fh.stdout.readlines()
arr.reverse()

res_map = {}
patten = re.compile(r'(.+) \[Debug\]: HandAlgorithmTask::execute statisc camera\[(\d+)\] face_body_count\[(\d+)\] process face\[(\d+)\]   recv_face_feature\[(\d+)\]; process body\[(\d+)\] recv_body_feature\[(\d+)\]');
for line in arr:
    content = line.decode('ascii')
    res = patten.match(content)
    if res:
        camera_id = res.group(2)
        if camera_id in res_map:
            continue

        cur_time = res.group(1)
        face_body_count = int(res.group(3))
        process_face = int(res.group(4))
        recv_face_fature = int(res.group(5))
        process_body = int(res.group(6))
        recv_body_feature = int(res.group(7))
        temp = {}
        temp['最后日志时间'] = cur_time
        temp['算法输出人脸和身体结构体(包含没有特征的)'] = face_body_count
        temp['c++处理过带有特征的人脸'] = process_face
        temp['算法输出的带有特征的人脸'] = recv_face_fature
        temp['c++处理的带有特征的身体'] = process_body
        temp['算法输出的带有特征的身体'] = recv_body_feature
        res_map[camera_id] = temp

wb = Workbook()
ws = wb.active
ws.title = '每路摄像头产量明细'
ws.sheet_properties.tabColor = "1072BA"

titles_labels = ['摄像头id','算法输出人脸和身体结构体(包含没有特征的)','算法输出的带有特征的人脸','c++处理过带有特征的人脸','算法输出的带有特征的身体','c++处理的带有特征的身体','最后日志时间']
for index in range(0, len(titles_labels)):
    ws.cell(row=1, column= index + 1, value=titles_labels[index])

print('摄像头数量 ', len(res_map))
for key in res_map:
    cur_row = ws.max_row + 1
    value_map = res_map[key]
    ws.cell(row= cur_row, column=1, value=key)
    print(key, res_map[key])
    for index in range(1, len(titles_labels)):
        ws.cell(row=cur_row, column=index + 1, value=value_map[titles_labels[index]])

wb.save("test.xlsx")