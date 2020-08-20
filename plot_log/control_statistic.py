#!/usr/bin/python3.5
# -*- coding: UTF-8 -*-
import re
from datetime import datetime
from openpyxl import Workbook
from clickhouse_driver import Client

def static_data(file_name):
    out_data = {}
    with open(file_name) as f:
        iter_f = iter(f)
        start_pt = datetime.now()
        end_pt = datetime.now()
        is_first = True
        for line in iter_f:
            patten = re.compile(r'(.+) \[Debug\]: SnapAlarmMongoDao::syncData2Java time statis camera_id: (\d+) type:\d person_id: (.+)$')
            match = patten.match(line)
            if not match:
                continue

            time_axis = datetime.strptime(match.group(1), '%Y-%m-%d %H:%M:%S.%f')
            if is_first:
                start_pt = time_axis
                is_first = False

            end_pt = time_axis

            camera_id = int(match.group(2))
            person_id = match.group(3)
            if camera_id in out_data:
                person_map = out_data[camera_id]
                if person_id in person_map:
                    out_data[camera_id][person_id] = out_data[camera_id][person_id] + 1
                else:
                    out_data[camera_id][person_id] = 1
            else:
                out_data[camera_id] = {person_id: 1}

    return start_pt, end_pt, out_data

#tuple 0-camera_id 1-start_pt 2-end_pt
def get_data_from_ck(params):
    client = Client(host='192.168.2.97', port='9000')
    sql = 'select person_id, count() as count from KoalaSnapData.snap_alarm where ts between {} and {} and camera_id={} group by person_id'.format(params[1], params[2], params[0])
    return client.execute(sql)

def get_person_count(person_id, input):
    for value in input:
        if value[0] == person_id:
            return value[1]

def save_detail(start_timestamp, end_timestamp, out_data, ws):
    ws.cell(row=1,column=1,value='camera_id')
    ws.cell(row=1, column=2, value='person_id')
    ws.cell(row=1, column=3, value='算法输出')
    ws.cell(row=1, column=4, value='ck数据库输出')

    row_num = 2
    for camera_id in out_data:
        camera_map = out_data[camera_id]
        ck_data = get_data_from_ck((camera_id, start_timestamp, end_timestamp))
        for person_id in camera_map:
            ws.cell(row=row_num, column=1, value=camera_id)
            ws.cell(row=row_num, column=2, value=int(person_id))
            ws.cell(row=row_num, column=3, value=camera_map[person_id])
            ws.cell(row=row_num, column=4, value=get_person_count(person_id, ck_data))
            row_num = row_num+1

def save_count(start_timestamp, end_timestamp, out_data, ws):
    ws.cell(row=1,column=1,value='camera_id')
    ws.cell(row=1, column=2, value='算法输出')
    ws.cell(row=1, column=3, value='ck数据库输出')

    for row_num, camera_id in enumerate(out_data, start=2):
        ws.cell(row=row_num, column=1, value=camera_id)
        ws.cell(row=row_num, column=2, value=len(out_data[camera_id]))
        ck_data = get_data_from_ck((camera_id, start_timestamp, end_timestamp))
        ws.cell(row=row_num, column=3, value=len(ck_data))


if __name__ == '__main__':
    start_pt, end_pt, out_data = static_data('/home/whl/seye-logs/KLMediaServer.log')
    print('start time is {}, end time is {}'.format(start_pt.strftime('%Y-%m-%d %H:%M:%S.%f'), end_pt.strftime('%Y-%m-%d %H:%M:%S.%f')))
    start_timestamp = int(start_pt.timestamp() * 1000)
    end_timestamp = int(end_pt.timestamp() * 1000)
    print('start stamp is {}, end stamp is {}'.format(start_timestamp, end_timestamp))

    wb = Workbook()
    sheet = wb.active
    sheet.title='deatil'
    save_detail(start_timestamp,end_timestamp,out_data, sheet)

    sheet = wb.create_sheet('all')
    save_count(start_timestamp, end_timestamp, out_data, sheet)

    wb.save("control.xlsx")