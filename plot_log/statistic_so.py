#!/usr/bin/python3.5
# -*- coding: UTF-8 -*-
import os
import re
from time import *
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook

def findlog_by_time_indir(log_dir, pattern, out_file_name):
    ofile = open(out_file_name, 'w')
    file_names = os.listdir(log_dir)
    for file_name in file_names:
        print(file_name)
        if not os.path.isdir(file_name):
            f = open(log_dir + "/" + file_name, "r")
            iter_f = iter(f)
            for line in iter_f:
                match = pattern.match(line)
                if match:
                    print(match.group())
                    ofile.writelines(match.group())
                    ofile.write("\n")

def statistics_count(input_file, pattmap, out_count):
    with open(input_file) as f:
        iter_f = iter(f)
        for line in iter_f:
            for key, value in pattmap.items():
                pattern = re.compile(r'' + value)
                match = pattern.match(line)
                if match:
                    camera_id = match.group(1)
                    if camera_id in out_count:
                        if key in out_count[camera_id]:
                            old_value = out_count[camera_id][key]
                            out_count[camera_id][key] = old_value + 1;
                        else:
                            out_count[camera_id][key] = 1
                    else:
                        out_count[camera_id]={key: 1}

def show_time_count(out_count, out_file, titles):
    wb = Workbook()
    ws = wb.active
    ws.title='deatil'
    titl_map = {}
    ws.cell(row=1,column=1,value='camera_id')
    for index, t in enumerate(titles, 2):
        titl_map[t] = index
        ws.cell(row=1, column=index, value=t)

    for row, camera_id in enumerate(out_count, 2):
        ws.cell(row=row,column=1,value=camera_id)
        for table in titles:
            if table in out_count[camera_id]:
                ws.cell(row=row,column=titl_map[table],value=out_count[camera_id][table])

    wb.save("statistic.xlsx")

def main():
    start_time = time()
    #findlog_by_time_indir("log", re.compile(r'2020-03-012 0[0-1]:\d\d.+$'), "log_bytime.txt")
    #SnapSceneMongoDao::syncSceneDataToJava time camera_id: 99 record id:43e35288967b11eaaa42e0d55e4aee45
    labes_map = {
        'SnapFace': '.+ \[Debug\]: SnapFaceMongoDao::syncData2Java camera_id: (\d+)$',
        'SnapBody': '.+ \[Debug\]: SnapBodyMongoDao::syncData2Java camera_id:(\d+) id:',
        'SnapScene': '.+ \[Debug\]: SnapSceneMongoDao::syncSceneDataToJava time camera_id: (\d+) record id:',
        'SnapAlarm': '.+ \[Debug\]: SnapAlarmMongoDao::syncData2Java time statis camera_id: (\d+) type:(?:\d+)',
        'Attribute': '.+ \[Debug\]: PersonAttributeMongoDao::syncData2Java time statis camera_id: (\d+)$',
        'RelationShip': '.+ \[Debug\]: RelationShipMongoDao::syncData2Java camera_id:(\d+)$'
    }

    out_count = {}
    print ('statistics start')
    statistics_count('/home/whl/seye-logs/KLMediaServer.log', labes_map, out_count)
    print ('statistics cost ', time() - start_time)

    show_time_count(out_count, 'result.xlsx', labes_map.keys())

if __name__ == '__main__':
    main()